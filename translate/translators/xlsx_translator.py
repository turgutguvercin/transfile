# XLSX Translation Module
import json
import re
import time
from pathlib import Path
from openpyxl import load_workbook
from openai import OpenAIError
from .utils import batched, client, MODEL


def _tool_schema():
    return [{
        "type": "function",
        "function": {
            "name": "return_translations",
            "description": "Return translations aligned with input indexes",
            "parameters": {
                "type": "object",
                "properties": {
                    "items": {
                        "type": "array",
                        "items": {
                            "type": "object",
                            "properties": {
                                "i": {"type": "integer"},
                                "t": {"type": "string"}
                            },
                            "required": ["i", "t"]
                        }
                    }
                },
                "required": ["items"]
            }
        }
    }]


def _build_prompt(items, src: str, tgt: str, note: str = "") -> str:
    intro = (
        f"You are a professional translator. Translate from {src} to {tgt}. "
        f"Preserve meaning, tone, punctuation, numbers, and line breaks. "
        f"Translate text exactly as given without adding commentary."
    )
    constraints = (
        "Return results ONLY by calling the function return_translations with an array 'items' "
        "of objects {i, t} where 'i' is the provided index and 't' is the translation. "
        "Return translations for ALL provided indexes, no more, no fewer, and do not reorder."
    )
    payload = json.dumps({"items": items}, ensure_ascii=False)
    extra = f"\n\nNote: {note}" if note else ""
    return f"{intro}\n{constraints}{extra}\n\nInput JSON:\n{payload}"


def _extract_items_from_response(r):
    choice = r.choices[0].message
    tool_calls = getattr(choice, "tool_calls", None)
    if tool_calls:
        for tc in tool_calls:
            try:
                if getattr(tc, "type", "") == "function" and getattr(tc.function, "name", "") == "return_translations":
                    args = tc.function.arguments
                    data = json.loads(args) if isinstance(args, str) else args
                    return data.get("items")
            except Exception:
                continue
    content = (choice.content or "").strip()
    if content:
        try:
            data = json.loads(content)
            if isinstance(data, dict) and "items" in data:
                return data["items"]
            if isinstance(data, list):
                return [{"i": i, "t": s} for i, s in enumerate(data)]
        except Exception:
            pass
    return None


_SYMBOLS_RE = re.compile(r"^[\W_]+$", re.UNICODE)


def _is_passthrough(text: str) -> bool:
    if text is None:
        return True
    s = str(text)
    if not s.strip():
        return True
    try:
        float(s.replace(",", "").replace(" ", ""))
        return True
    except Exception:
        pass
    return bool(_SYMBOLS_RE.match(s))


def translate_xlsx_chunk(texts, src, tgt, max_attempts: int = 3):
    if not texts:
        return []

    n = len(texts)
    remaining_indexes = list(range(n))
    collected = {}

    for i, val in enumerate(texts):
        if _is_passthrough(val):
            collected[i] = str(val)

    full_items = [{"i": i, "s": texts[i]} for i in range(n) if i not in collected]

    attempts = 0
    while attempts < max_attempts and remaining_indexes:
        attempts += 1
        ask_items = (
            full_items if attempts == 1 else [{"i": i, "s": texts[i]} for i in remaining_indexes]
        )
        note = "" if attempts == 1 else (
            f"Retry {attempts-1}: Only return translations for the listed missing indexes. "
            f"Do not include any other indexes."
        )
        try:
            r = client.chat.completions.create(
                model=MODEL,
                temperature=0.1,
                messages=[{
                    "role": "user",
                    "content": _build_prompt(ask_items, src, tgt, note)
                }],
                tools=_tool_schema(),
                tool_choice={"type": "function", "function": {"name": "return_translations"}},
            )
        except OpenAIError as e:
            raise RuntimeError(f"I am really sorry an error happened/ Çok ama çok üzgünüm bir hata oluştu {e}")
        except Exception as e:
            raise RuntimeError(f"Your app run into a problem :( {e} ")

        items = _extract_items_from_response(r)
        if not items:
            time.sleep(0.2)
            continue

        for obj in items:
            try:
                i = int(obj.get("i"))
                t = obj.get("t")
            except Exception:
                continue
            if i in range(n) and isinstance(t, str):
                collected[i] = t

        remaining_indexes = [i for i in range(n) if i not in collected]
        if remaining_indexes:
            time.sleep(0.2)

    if len(collected) != n:
        missing = [i for i in range(n) if i not in collected]
        for i in missing:
            collected[i] = str(texts[i])

    return [collected[i] for i in range(n)]

def translate_xlsx(in_path: Path, out_path: Path, src: str, tgt: str, batch_size: int = 80):
    """Translate an XLSX file"""
    wb = load_workbook(str(in_path))
    total_cells, changed = 0, 0

    for ws in wb.worksheets:
        coords, texts = [], []
        for row in ws.iter_rows():
            for cell in row:
                cell_value = cell.value
                # Translate only non-empty strings; skip formulas
                if isinstance(cell_value, str) and cell_value.strip() and not cell_value.strip().startswith("="):
                    coords.append((ws.title, cell.row, cell.column))
                    texts.append(cell_value)
        total_cells += len(coords)

        idx = 0
        for chunk in batched(texts, batch_size):
            translated_chunk = translate_xlsx_chunk(chunk, src, tgt)
            for t in translated_chunk:
                _, r, col = coords[idx]
                ws.cell(r, col).value = t
                idx += 1
                changed += 1

    wb.save(str(out_path))
    print(f"Translated XLSX ({changed}/{total_cells}) → {out_path}")
